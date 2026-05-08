from openpyxl import load_workbook

from app.schemas.parser import (
    ParsedDocument,
    ParsedPage,
    ParsedSection,
    ParsedTable,
)


class XLSXParser:

    @staticmethod
    def parse(
        file_path: str,
    ) -> ParsedDocument:

        try:

            workbook = load_workbook(
                filename=file_path,
                data_only=True,
            )

            parsed_pages = []
            parsed_sections = []
            parsed_tables = []

            page_number = 1

            for sheet in workbook.worksheets:

                rows = list(
                    sheet.iter_rows(
                        values_only=True
                    )
                )

                if not rows:
                    continue

                headers = [
                    str(cell).strip()
                    if cell is not None
                    else ""
                    for cell in rows[0]
                ]

                table_rows = []

                for row in rows[1:]:

                    cleaned_row = [
                        str(cell).strip()
                        if cell is not None
                        else ""
                        for cell in row
                    ]

                    table_rows.append(cleaned_row)

                parsed_tables.append(
                    ParsedTable(
                        title=sheet.title,
                        headers=headers,
                        rows=table_rows,
                        page_number=page_number,
                    )
                )

                section_content = "\n".join(
                    [
                        " | ".join(headers),
                        *[
                            " | ".join(row)
                            for row in table_rows
                        ],
                    ]
                )

                parsed_sections.append(
                    ParsedSection(
                        title=sheet.title,
                        content=section_content,
                        page_number=page_number,
                    )
                )

                parsed_pages.append(
                    ParsedPage(
                        page_number=page_number,
                        text=section_content,
                    )
                )

                page_number += 1

            return ParsedDocument(
                sections=parsed_sections,
                tables=parsed_tables,
                pages=parsed_pages,
                total_pages=len(parsed_pages),
            )

        except Exception as exc:
            raise ValueError(
                f"Excel parsing failed: {exc}"
            )